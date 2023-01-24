import yadisk
import openpyxl
from openpyxl import load_workbook

y = yadisk.YaDisk(token="y0_AgAAAABkLI6WAAhdVQAAAADUKP-n_QegnmwnQTSiQj9SNN02gziNKW4")
y.download('/123.xlsx', "123.xlsx")  # cкачивание файла
wb = openpyxl.load_workbook(filename="123.xlsx")
wb.active = 0
sheetVlad = wb.active
wb.close()

def skach():
    y.download('/123.xlsx', "123.xlsx")  # cкачивание файла

def obnovl():
    wb = openpyxl.load_workbook(filename="123.xlsx")
    wb.active = 0
    sheetVlad = wb.active


def nom_counter():
    a2 = (sheetVlad['A2'].value)   #0
    a3 = (sheetVlad['A3'].value)   #1
    a4 = (sheetVlad['A4'].value)   #2
    a5 = (sheetVlad['A5'].value)   #3
    a6 = (sheetVlad['A6'].value)   #4
    a7 = (sheetVlad['A7'].value)   #5
    a8 = (sheetVlad['A8'].value)   #6
    a9 = (sheetVlad['A9'].value)   #7
    a10 = (sheetVlad['A10'].value) #8
    a11 = (sheetVlad['A11'].value) #9
    a12 = (sheetVlad['A12'].value) #10
    a13 = (sheetVlad['A13'].value) #11
    a14 = (sheetVlad['A14'].value) #12
    a15 = (sheetVlad['A15'].value) #13
    a16 = (sheetVlad['A16'].value) #14
    a17 = (sheetVlad['A17'].value) #15
    a18 = (sheetVlad['A18'].value) #16
    a19 = (sheetVlad['A19'].value) #17
    a20 = (sheetVlad['A20'].value) #18
    a21 = (sheetVlad['A21'].value)  # 19
    a22 = (sheetVlad['A22'].value)  # 18
    a23 = (sheetVlad['A23'].value)  # 18
    a24 = (sheetVlad['A24'].value)  # 18
    a25 = (sheetVlad['A25'].value)  # 18
    a26 = (sheetVlad['A26'].value)  # 18
    a27 = (sheetVlad['A27'].value)  # 18
    a28 = (sheetVlad['A28'].value)  # 18
    a29 = (sheetVlad['A29'].value)  # 18
    a30 = (sheetVlad['A30'].value)  # 18
    a31 = (sheetVlad['A31'].value)  # 18
    a32 = (sheetVlad['A32'].value)  # 18
    a33 = (sheetVlad['A33'].value)  # 18
    a34 = (sheetVlad['A34'].value)  # 18
    a35 = (sheetVlad['A35'].value)  # 18
    a36 = (sheetVlad['A36'].value)  # 18
    a37 = (sheetVlad['A37'].value)  # 18
    a38 = (sheetVlad['A38'].value)  # 18
    a39 = (sheetVlad['A39'].value)  # 18
    a40 = (sheetVlad['A40'].value)  # 18
    a41 = (sheetVlad['A41'].value)  # 18
    a42 = (sheetVlad['A42'].value)  # 18
    a43 = (sheetVlad['A43'].value)  # 18
    a44 = (sheetVlad['A44'].value)  # 18
    a45 = (sheetVlad['A45'].value)  # 18
    a46 = (sheetVlad['A46'].value)  # 18
    a47 = (sheetVlad['A47'].value)  # 18
    a48 = (sheetVlad['A48'].value)  # 18
    a49 = (sheetVlad['A49'].value)  # 18
    a50 = (sheetVlad['A50'].value)  # 18
    a51 = (sheetVlad['A51'].value)  # 18
    a52 = (sheetVlad['A52'].value)  # 18
    a53 = (sheetVlad['A53'].value)  # 18
    a54 = (sheetVlad['A54'].value)  # 18
    a55 = (sheetVlad['A55'].value)  # 18
    a56 = (sheetVlad['A56'].value)  # 18
    a57 = (sheetVlad['A57'].value)  # 18
    a58 = (sheetVlad['A58'].value)  # 18
    a59 = (sheetVlad['A59'].value)  # 18
    a60 = (sheetVlad['A60'].value)  # 18
    a61 = (sheetVlad['A61'].value)  # 18
    a62 = (sheetVlad['A62'].value)  # 18
    a63 = (sheetVlad['A63'].value)  # 18
    a64 = (sheetVlad['A64'].value)  # 18
    a65 = (sheetVlad['A65'].value)  # 18
    a66 = (sheetVlad['A66'].value)  # 18
    a67 = (sheetVlad['A67'].value)  # 18
    a68 = (sheetVlad['A68'].value)  # 18
    a69 = (sheetVlad['A69'].value)  # 18
    a70 = (sheetVlad['A70'].value)  # 18
    a71 = (sheetVlad['A71'].value)  # 18
    a72 = (sheetVlad['A72'].value)  # 18
    a73 = (sheetVlad['A73'].value)  # 18
    a74 = (sheetVlad['A74'].value)  # 18
    a75 = (sheetVlad['A75'].value)  # 18
    a76 = (sheetVlad['A76'].value)  # 18
    a77 = (sheetVlad['A77'].value)  # 18
    a78 = (sheetVlad['A78'].value)  # 18
    a79 = (sheetVlad['A79'].value)  # 18
    a80 = (sheetVlad['A80'].value)  # 18
    a81 = (sheetVlad['A81'].value)  # 18
    a82 = (sheetVlad['A82'].value)  # 18
    a83 = (sheetVlad['A83'].value)  # 18
    a84 = (sheetVlad['A84'].value)  # 18
    a85 = (sheetVlad['A85'].value)  # 18
    a86 = (sheetVlad['A86'].value)  # 18
    a87 = (sheetVlad['A87'].value)  # 18
    a88 = (sheetVlad['A88'].value)  # 18
    a89 = (sheetVlad['A89'].value)  # 18
    a90 = (sheetVlad['A90'].value)  # 18
    a91 = (sheetVlad['A91'].value)  # 18
    a92 = (sheetVlad['A92'].value)  # 18
    a93 = (sheetVlad['A93'].value)  # 18
    a94 = (sheetVlad['A94'].value)  # 18
    a95 = (sheetVlad['A95'].value)  # 18
    a96 = (sheetVlad['A96'].value)  # 18
    a97 = (sheetVlad['A97'].value)  # 18
    a98 = (sheetVlad['A98'].value)  # 18
    a99 = (sheetVlad['A99'].value)  # 18
    a100 = (sheetVlad['A100'].value)  # 18
    a101 = (sheetVlad['A101'].value)
    a102 = (sheetVlad['A102'].value)
    a103 = (sheetVlad['A103'].value)
    a104 = (sheetVlad['A104'].value)
    a105 = (sheetVlad['A105'].value)
    a106 = (sheetVlad['A106'].value)
    a107 = (sheetVlad['A107'].value)
    a108 = (sheetVlad['A108'].value)
    a109 = (sheetVlad['A109'].value)
    a110 = (sheetVlad['A110'].value)
    a111 = (sheetVlad['A111'].value)
    a112 = (sheetVlad['A112'].value)
    a113 = (sheetVlad['A113'].value)
    a114 = (sheetVlad['A114'].value)
    a115 = (sheetVlad['A115'].value)
    a116 = (sheetVlad['A116'].value)
    a117 = (sheetVlad['A117'].value)
    a118 = (sheetVlad['A118'].value)
    a119 = (sheetVlad['A119'].value)
    a120 = (sheetVlad['A120'].value)
    a121 = (sheetVlad['A121'].value)
    a122 = (sheetVlad['A122'].value)
    a123 = (sheetVlad['A123'].value)
    a124 = (sheetVlad['A124'].value)
    a125 = (sheetVlad['A125'].value)
    a126 = (sheetVlad['A126'].value)
    a127 = (sheetVlad['A127'].value)
    a128 = (sheetVlad['A128'].value)
    a129 = (sheetVlad['A129'].value)
    a130 = (sheetVlad['A130'].value)
    a131 = (sheetVlad['A131'].value)
    a132 = (sheetVlad['A132'].value)
    a133 = (sheetVlad['A133'].value)
    a134 = (sheetVlad['A134'].value)
    a135 = (sheetVlad['A135'].value)
    a136 = (sheetVlad['A136'].value)
    a137 = (sheetVlad['A137'].value)
    a138 = (sheetVlad['A138'].value)
    a139 = (sheetVlad['A139'].value)
    a140 = (sheetVlad['A140'].value)
    a141 = (sheetVlad['A141'].value)
    a142 = (sheetVlad['A142'].value)
    a143 = (sheetVlad['A143'].value)
    a144 = (sheetVlad['A144'].value)
    a145 = (sheetVlad['A145'].value)
    a146 = (sheetVlad['A146'].value)
    a147 = (sheetVlad['A147'].value)
    a148 = (sheetVlad['A148'].value)
    a149 = (sheetVlad['A149'].value)
    a150 = (sheetVlad['A150'].value)

    return [a2, a3, a4, a5, a6, a7, a8, a9, a10, #9
            a11,a12,a13,a14,a15,a16,a17,a18,a19,a20, #10
            a21,a22,a23,a24,a25,a26,a27,a28,a29,a30,
            a31,a32,a33,a34,a35,a36,a37,a38,a39,a40,
            a41,a42,a43,a44,a45,a46,a47,a48,a49,a50,
            a51,a52,a53,a54,a55,a56,a57,a58,a59,a60,
            a61,a62,a63,a64,a65,a66,a67,a68,a69,a70,
            a71,a72,a73,a74,a75,a76,a77,a78,a79,a80,
            a81,a82,a83,a84,a85,a86,a87,a88,a89,a90,
            a91,a92,a93,a94,a95,a96,a97,a98,a99,a100, #90
            a101,a102,a103,a104,a105,a106,a107,a108,a109,a110, #50
            a111,a112,a113,a114,a115,a116,a117,a118,a119,a120,
            a121,a122,a123,a124,a125,a126,a127,a128,a129,a130,
            a131,a132,a133,a134,a135,a136,a137,a138,a139,a140,
            a141,a142,a143,a144,a145,a146,a147,a148,a149,a150]


def nom_name():
    wb = openpyxl.load_workbook(filename="123.xlsx")
    wb.active = 0
    sheetVlad = wb.active
    b2 = (sheetVlad['B2'].value)
    b3 = (sheetVlad['B3'].value)
    b4 = (sheetVlad['B4'].value)
    b5 = (sheetVlad['B5'].value)
    b6 = (sheetVlad['B6'].value)
    b7 = (sheetVlad['B7'].value)
    b8 = (sheetVlad['B8'].value)
    b9 = (sheetVlad['B9'].value)
    b10 = (sheetVlad['B10'].value)
    b11 = (sheetVlad['B11'].value)
    b12 = (sheetVlad['B12'].value)
    b13 = (sheetVlad['B13'].value)
    b14 = (sheetVlad['B14'].value)
    b15 = (sheetVlad['B15'].value)
    b16 = (sheetVlad['B16'].value)
    b17 = (sheetVlad['B17'].value)
    b18 = (sheetVlad['B18'].value)
    b19 = (sheetVlad['B19'].value)
    b20 = (sheetVlad['B20'].value)
    b21 = (sheetVlad['B21'].value)  # 18
    b22 = (sheetVlad['B22'].value)  # 18
    b23 = (sheetVlad['B23'].value)  # 18
    b24 = (sheetVlad['B24'].value)  # 18
    b25 = (sheetVlad['B25'].value)  # 18
    b26 = (sheetVlad['B26'].value)  # 18
    b27 = (sheetVlad['B27'].value)  # 18
    b28 = (sheetVlad['B28'].value)  # 18
    b29 = (sheetVlad['B29'].value)  # 18
    b30 = (sheetVlad['B30'].value)  # 18
    b31 = (sheetVlad['B31'].value)  # 18
    b32 = (sheetVlad['B32'].value)  # 18
    b33 = (sheetVlad['B33'].value)  # 18
    b34 = (sheetVlad['B34'].value)  # 18
    b35 = (sheetVlad['B35'].value)  # 18
    b36 = (sheetVlad['B36'].value)  # 18
    b37 = (sheetVlad['B37'].value)  # 18
    b38 = (sheetVlad['B38'].value)  # 18
    b39 = (sheetVlad['B39'].value)  # 18
    b40 = (sheetVlad['B40'].value)  # 18
    b41 = (sheetVlad['B41'].value)  # 18
    b42 = (sheetVlad['B42'].value)  # 18
    b43 = (sheetVlad['B43'].value)  # 18
    b44 = (sheetVlad['B44'].value)  # 18
    b45 = (sheetVlad['B45'].value)  # 18
    b46 = (sheetVlad['B46'].value)  # 18
    b47 = (sheetVlad['B47'].value)  # 18
    b48 = (sheetVlad['B48'].value)  # 18
    b49 = (sheetVlad['B49'].value)  # 18
    b50 = (sheetVlad['B50'].value)  # 18
    b51 = (sheetVlbd['B51'].vblue)  # 18
    b52 = (sheetVlbd['B52'].vblue)  # 18
    b53 = (sheetVlbd['B53'].vblue)  # 18
    b54 = (sheetVlbd['B54'].vblue)  # 18
    b55 = (sheetVlbd['B55'].vblue)  # 18
    b56 = (sheetVlbd['B56'].vblue)  # 18
    b57 = (sheetVlbd['B57'].vblue)  # 18
    b58 = (sheetVlbd['B58'].vblue)  # 18
    b59 = (sheetVlbd['B59'].vblue)  # 18
    b60 = (sheetVlbd['B60'].vblue)  # 18
    b61 = (sheetVlbd['B61'].vblue)  # 18
    b62 = (sheetVlbd['B62'].vblue)  # 18
    b63 = (sheetVlbd['B63'].vblue)  # 18
    b64 = (sheetVlbd['B64'].vblue)  # 18
    b65 = (sheetVlbd['B65'].vblue)  # 18
    b66 = (sheetVlbd['B66'].vblue)  # 18
    b67 = (sheetVlbd['B67'].vblue)  # 18
    b68 = (sheetVlbd['B68'].vblue)  # 18
    b69 = (sheetVlbd['B69'].vblue)  # 18
    b70 = (sheetVlbd['B70'].vblue)  # 18
    b71 = (sheetVlbd['B71'].vblue)  # 18
    b72 = (sheetVlbd['B72'].vblue)  # 18
    b73 = (sheetVlbd['B73'].vblue)  # 18
    b74 = (sheetVlbd['B74'].vblue)  # 18
    b75 = (sheetVlbd['B75'].vblue)  # 18
    b76 = (sheetVlbd['B76'].vblue)  # 18
    b77 = (sheetVlbd['B77'].vblue)  # 18
    b78 = (sheetVlbd['B78'].vblue)  # 18
    b79 = (sheetVlbd['B79'].vblue)  # 18
    b80 = (sheetVlbd['B80'].vblue)  # 18
    b81 = (sheetVlbd['B81'].vblue)  # 18
    b82 = (sheetVlbd['B82'].vblue)  # 18
    b83 = (sheetVlbd['B83'].vblue)  # 18
    b84 = (sheetVlbd['B84'].vblue)  # 18
    b85 = (sheetVlbd['B85'].vblue)  # 18
    b86 = (sheetVlbd['B86'].vblue)  # 18
    b87 = (sheetVlbd['B87'].vblue)  # 18
    b88 = (sheetVlbd['B88'].vblue)  # 18
    b89 = (sheetVlbd['B89'].vblue)  # 18
    b90 = (sheetVlbd['B90'].vblue)  # 18
    b91 = (sheetVlbd['B91'].vblue)  # 18
    b92 = (sheetVlbd['B92'].vblue)  # 18
    b93 = (sheetVlbd['B93'].vblue)  # 18
    b94 = (sheetVlbd['B94'].vblue)  # 18
    b95 = (sheetVlbd['B95'].vblue)  # 18
    b96 = (sheetVlbd['B96'].vblue)  # 18
    b97 = (sheetVlbd['B97'].vblue)  # 18
    b98 = (sheetVlbd['B98'].vblue)  # 18
    b99 = (sheetVlbd['B99'].vblue)  # 18
    b100 = (sheetVlbd['B100'].vblue)  # 18
    b101 = (sheetVlbd['B101'].vblue)
    b102 = (sheetVlbd['B102'].vblue)
    b103 = (sheetVlbd['B103'].vblue)
    b104 = (sheetVlbd['B104'].vblue)
    b105 = (sheetVlbd['B105'].vblue)
    b106 = (sheetVlbd['B106'].vblue)
    b107 = (sheetVlbd['B107'].vblue)
    b108 = (sheetVlbd['B108'].vblue)
    b109 = (sheetVlbd['B109'].vblue)
    b110 = (sheetVlbd['B110'].vblue)
    b111 = (sheetVlbd['B111'].vblue)
    b112 = (sheetVlbd['B112'].vblue)
    b113 = (sheetVlbd['B113'].vblue)
    b114 = (sheetVlbd['B114'].vblue)
    b115 = (sheetVlbd['B115'].vblue)
    b116 = (sheetVlbd['B116'].vblue)
    b117 = (sheetVlbd['B117'].vblue)
    b118 = (sheetVlbd['B118'].vblue)
    b119 = (sheetVlbd['B119'].vblue)
    b120 = (sheetVlbd['B120'].vblue)
    b121 = (sheetVlbd['B121'].vblue)
    b122 = (sheetVlbd['B122'].vblue)
    b123 = (sheetVlbd['B123'].vblue)
    b124 = (sheetVlbd['B124'].vblue)
    b125 = (sheetVlbd['B125'].vblue)
    b126 = (sheetVlbd['B126'].vblue)
    b127 = (sheetVlbd['B127'].vblue)
    b128 = (sheetVlbd['B128'].vblue)
    b129 = (sheetVlbd['B129'].vblue)
    b130 = (sheetVlbd['B130'].vblue)
    b131 = (sheetVlbd['B131'].vblue)
    b132 = (sheetVlbd['B132'].vblue)
    b133 = (sheetVlbd['B133'].vblue)
    b134 = (sheetVlbd['B134'].vblue)
    b135 = (sheetVlbd['B135'].vblue)
    b136 = (sheetVlbd['B136'].vblue)
    b137 = (sheetVlbd['B137'].vblue)
    b138 = (sheetVlbd['B138'].vblue)
    b139 = (sheetVlbd['B139'].vblue)
    b140 = (sheetVlbd['B140'].vblue)
    b141 = (sheetVlbd['B141'].vblue)
    b142 = (sheetVlbd['B142'].vblue)
    b143 = (sheetVlbd['B143'].vblue)
    b144 = (sheetVlbd['B144'].vblue)
    b145 = (sheetVlbd['B145'].vblue)
    b146 = (sheetVlbd['B146'].vblue)
    b147 = (sheetVlbd['B147'].vblue)
    b148 = (sheetVlbd['B148'].vblue)
    b149 = (sheetVlbd['B149'].vblue)
    b150 = (sheetVlbd['B150'].vblue)

    return [b2, b3, b4, b5, b6, b7, b8, b9, b10,  # 9
            b11, b12, b13, b14, b15, b16, b17, b18, b19, b20,  # 10
            b21, b22, b23, b24, b25, b26, b27, b28, b29, b30,
            b31, b32, b33, b34, b35, b36, b37, b38, b39, b40,
            b41, b42, b43, b44, b45, b46, b47, b48, b49, b50,
            b51, b52, b53, b54, b55, b56, b57, b58, b59, b60,
            b61, b62, b63, b64, b65, b66, b67, b68, b69, b70,
            b71, b72, b73, b74, b75, b76, b77, b78, b79, b80,
            b81, b82, b83, b84, b85, b86, b87, b88, b89, b90,
            b91, b92, b93, b94, b95, b96, b97, b98, b99, b100,  # 90
            b101, b102, b103, b104, b105, b106, b107, b108, b109, b110,  # 50
            b111, b112, b113, b114, b115, b116, b117, b118, b119, b120,
            b121, b122, b123, b124, b125, b126, b127, b128, b129, b130,
            b131, b132, b133, b134, b135, b136, b137, b138, b139, b140,
            b141, b142, b143, b144, b145, b146, b147, b148, b149, b150]







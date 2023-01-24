import openpyxl
import yadisk
y = yadisk.YaDisk(token="y0_AgAAAABkLKNnAAhdVQAAAADOP4ODkm5Bu17jRHaSlteu5qh343W0KpY")
y.download('/Tsena.xlsx', "Tsena.xlsx")  # cкачивание файла
wb = openpyxl.load_workbook(filename="Tsena.xlsx")
# print(wb.sheetnames)
wb.active = 0
sheetVlad = wb.active

def TsVl():
    wb = openpyxl.load_workbook(filename="Tsena.xlsx")
    wb.active = 0
    sheetVlad = wb.active
    #шмаль
    s05 = (sheetVlad['B2'].value) #0
    s1 = (sheetVlad['B3'].value) #1
    s2 = (sheetVlad['B4'].value) #2
    s3 = (sheetVlad['B5'].value) #3
    s5 = (sheetVlad['B6'].value) #4

    #меф
    m05 = (sheetVlad['B8'].value)  # 5
    m1 = (sheetVlad['B9'].value)   # 6
    m2 = (sheetVlad['B10'].value)  # 7
    m3 = (sheetVlad['B11'].value)  # 8
    m5 = (sheetVlad['B12'].value)  # 9

    #PVP
    p03 = (sheetVlad['B14'].value)  # 10
    p05 = (sheetVlad['B15'].value)  # 11
    p1 = (sheetVlad['B16'].value)   # 12
    p2 = (sheetVlad['B17'].value)   # 13
    p3 = (sheetVlad['B18'].value)   # 14
    p5 = (sheetVlad['B19'].value)   # 15

    #мука
    mm05 = (sheetVlad['B21'].value) # 16
    mm1 = (sheetVlad['B22'].value)  # 17
    mm2 = (sheetVlad['B23'].value)  # 18
    mm3 = (sheetVlad['B24'].value)  # 19

    #смола
    sm1 = (sheetVlad['B26'].value)  # 20
    sm2 = (sheetVlad['B27'].value)  # 21

    # экстази
    e1 = (sheetVlad['B29'].value)  # 22
    e2 = (sheetVlad['B30'].value)  # 23

    return [s05, s1, s2, s3, s5,
            m05, m1, m2, m3, m5,
            p03, p05, p1, p2, p3, p5,
            mm05, mm1, mm2, mm3,
            sm1, sm2,
            e1, e2]

def TsAr():
    wb = openpyxl.load_workbook(filename="Tsena.xlsx")
    wb.active = 0
    sheetVlad = wb.active
    #шмаль
    as05 = (sheetVlad['C2'].value) #0
    as1 = (sheetVlad['C3'].value) #1
    as2 = (sheetVlad['C4'].value) #2
    as3 = (sheetVlad['C5'].value) #3
    as5 = (sheetVlad['C6'].value) #4

    #меф
    am05 = (sheetVlad['C8'].value)  # 5
    am1 = (sheetVlad['C9'].value)   # 6
    am2 = (sheetVlad['C10'].value)  # 7
    am3 = (sheetVlad['C11'].value)  # 8
    am5 = (sheetVlad['C12'].value)  # 9

    #PVP
    ap03 = (sheetVlad['C14'].value)  # 10
    ap05 = (sheetVlad['C15'].value)  # 11
    ap1 = (sheetVlad['C16'].value)   # 12
    ap2 = (sheetVlad['C17'].value)   # 13
    ap3 = (sheetVlad['C18'].value)   # 14
    ap5 = (sheetVlad['C19'].value)   # 15

    #мука
    amm05 = (sheetVlad['C21'].value) # 16
    amm1 = (sheetVlad['C22'].value)  # 17
    amm2 = (sheetVlad['C23'].value)  # 18
    amm3 = (sheetVlad['C24'].value)  # 19

    #смола
    asm1 = (sheetVlad['C26'].value)  # 20
    asm2 = (sheetVlad['C27'].value)  # 21

    # экстази
    ae1 = (sheetVlad['C29'].value)  # 22
    ae2 = (sheetVlad['C30'].value)  # 23

    return [as05, as1, as2, as3, as5,
            am05, am1, am2, am3, am5,
            ap03, ap05, ap1, ap2, ap3, ap5,
            amm05, amm1, amm2, amm3,
            asm1, asm2,
            ae1, ae2]